import os
import json
from openpyxl.utils import get_column_letter
import openpyxl
 
def mkdir(path):
 
	folder = os.path.exists(path)
 
	if not folder:
		print("---  new folder...  ---")
		os.makedirs(path)
		print("---  OK  ---")
 
	else:
		print("---  There is this folder!  ---")
		

def getresult_n(number):
	title1 = ['玩家','第K个玩家进场','此时的进场成本','进场后的盈利','该玩家达到不亏需要拉的人头数','该玩家不亏时总人口数','总盘缺口资金数','总盘缺口人头数','进来的总资金数','出去的总资金数','总流水','文章已打赏数（总参与人数）']
	title2 = ['每个玩家盈亏', '第K个玩家进场','玩家']
	workbook = openpyxl.Workbook()
	worksheet = workbook.active
	worksheet.cell(2,1,"作者")
	for i in range(number-1):
		worksheet.cell(i+3, 1,i+1)	# 玩家 col A
	for a in range(len(title1)):	# title raw 1
		worksheet.cell(1, a+1, title1[a])	
	for i in range(number):			
		worksheet.cell(i+2, 2, i)	# 第K个玩家进场 col B
		worksheet.cell(i+2, 3, i)	# 此时的进场成本 col C
		worksheet.cell(i+2, 4, (-i))	# 进场后的盈利 col D
		worksheet.cell(i+2,5,i)		# 该玩家达到不亏需要拉的人头数 col E
		worksheet.cell(i+2,6,2*i+1)	# 该玩家不亏时总人口数 col F
		if(i % 2 == 0):				# 总盘缺口资金数 G
			worksheet.cell(i+2,7,-i*(i+2)/4)
		else:
			worksheet.cell(i+2,7,-(i+1)*(i+1)/4)
		worksheet.cell(i+2,8,i)				# 总盘缺口人头数 H
		worksheet.cell(i+2,9,(1+i)*i/2)		# 进来的总资金数 I
		worksheet.cell(i+2,10,(1+i)*i/2)	# 出去的总资金数 J
		worksheet.cell(i+2,11,(1+i)*i)		# 总流水 K
		worksheet.cell(i+2,12,1+i)			# 总参与人数 L

	worksheet2 = workbook.create_sheet()
	worksheet2.cell(1,1,"每个玩家盈亏")
	worksheet2.cell(1,2,"第K个玩家进场")
	worksheet2.cell(2,1,'玩家')
	for j in range(number):
		worksheet2.cell(2,3+j,j)
	for k in range(number):
		worksheet2.cell(3+k, 2, k)

	for x in range(number):			# 代表列
		for y in range(number-x):		# 代表行
			worksheet2.cell(3+y+x,3+x,y-x)

	worksheet2.title = "每个玩家实时盈亏"
	# 调整行高
	column_widths = []
	for row in worksheet.rows:
		for i, cell in enumerate(row):
			if len(column_widths) > i:
				if len(str(cell.value)) > column_widths[i]:
					column_widths[i] = len(str(cell.value))
			else:
				column_widths += [len(cell.value)]

	for i, column_width in enumerate(column_widths):
		worksheet.column_dimensions[get_column_letter(i+1)].width = column_width

	column_widths = []
	for row in worksheet2.rows:
		for i, cell in enumerate(row):
			if len(column_widths) > i:
				if len(str(cell.value)) > column_widths[i]:
					column_widths[i] = len(str(cell.value))
			else:
				column_widths += [len(str(cell.value))]

	for i, column_width in enumerate(column_widths):
		worksheet2.column_dimensions[get_column_letter(i+1)].width = column_width

	filename = "etheremon"+str(number)+".xlsx"
	workbook.save(filename)

getresult_n(100)