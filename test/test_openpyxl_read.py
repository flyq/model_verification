import openpyxl

workbook = openpyxl.load_workbook("test_openpyxl.xlsx")

worksheet = workbook["Sheet1"]
# workbook[shenames[1]]

rows=worksheet.max_row
columns=worksheet.max_column

for row in worksheet.rows:
    for cell in row:
        print(cell.value,end=" ")
    print()
"""
各省市 工资性收入 家庭经营纯收入 财产性收入 转移性收入 食品 衣着 居住 家庭设备及服务 ……
北京市 5047.4 1957.1 678.8 592.2 1879.0 451.6 859.4 303.5 698.1 844.1 575.8 113.1 ……
天津市 3247.9 2707.4 126.4 146.3 1212.6 265.3 664.4 122.4 441.3 315.6 263.2 56.1 ……
……
"""
 
for col in worksheet.columns:
    for cell in col:
        print(cell.value,end=" ")
    print()
 
'''
各省市 北京市 天津市 河北省 山西省 内蒙古自治区 辽宁省 吉林省 黑龙江省 上海市 江苏省 浙江省 ……
工资性收入 5047.4 3247.9 1514.7 1374.3 590.7 1499.5 605.1 654.9 6686.0 3104.8 3575.1 ……
家庭经营纯收入 1957.1 2707.4 2039.6 1622.9 2406.2 2210.8 2556.7 2521.5 767.7 2271.4  ……
……
'''


#输出特定的行
for cell in list(worksheet.rows)[3]:  #获取第四行的数据
    print(cell.value,end=" ")
print()
#河北省 1514.7 2039.6 107.7 139.8 915.5 167.9 531.7 115.8 285.7 265.4 166.3 47.0
 
#输出特定的列
for cell in list(worksheet.columns)[2]:  #获取第三列的数据
    print(cell.value,end=" ")
print()
#家庭经营纯收入 1957.1 2707.4 2039.6 1622.9 2406.2 2210.8 2556.7 2521.5 767.7 2271.4 3084.3……
 
#已经转换成list类型，自然是从0开始计数。


for rows in list(worksheet.rows)[0:3]:
    for cell in rows[0:3]:
        print(cell.value,end=" ")
    print()
'''
各省市 工资性收入 家庭经营纯收入 
北京市 5047.4 1957.1 
天津市 3247.9 2707.4 
'''
 
for i in range(1, 4):
    for j in range(1, 4):
        print(worksheet.cell(row=i, column=j).value,end=" ")
    print()
'''
各省市 工资性收入 家庭经营纯收入 
北京市 5047.4 1957.1 
天津市 3247.9 2707.4 
'''

#精确读取表格中的某一单元格
content_A1= worksheet['A1'].value
print(content_A1)
 
content_A1=worksheet.cell(row=1,column=1).value
#等同于 content_A1=worksheet.cell(1,1).value
print(content_A1)
#此处的行数和列数都是从1开始计数的，而在xlrd中是由0开始计数的
